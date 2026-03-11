#!/usr/bin/env python3
"""
XLSX to PDF Quality Test Script
================================
Tests XLSX to PDF conversion quality, specifically addressing:
- Content visibility (no truncation)
- Visual fidelity (proper scaling)
- Complete page rendering

Usage:
    python tests/e2e/test_xlsx_quality.py
"""

import asyncio
import os
import sys
import tempfile
from pathlib import Path

# Add backend to path
sys.path.insert(0, "/Users/caolei/Desktop/文件处理全能助手/backend")

# Test configuration
OUTPUT_DIR = "/Users/caolei/Desktop/文件处理全能助手/test_screenshots/xlsx_quality"
XLSX_FILE = "/Users/caolei/Desktop/文件处理全能助手/test_samples/5.2025计算机学院团委学生会换届汇总表.xlsx"


def log_step(step: str, message: str, success: bool = True):
    """Log test step"""
    icon = "✓" if success else "✗"
    print(f"[{icon}] {step}: {message}")


def analyze_xlsx_dimensions(file_path: str) -> dict:
    """Analyze XLSX file to get visual dimensions"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column

    # Calculate content bounds
    min_col = max_col
    max_col_used = 1
    min_row = max_row
    max_row_used = 1

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                min_col = min(min_col, cell.column)
                max_col_used = max(max_col_used, cell.column)
                min_row = min(min_row, cell.row)
                max_row_used = max(max_row_used, cell.row)

    # Get column widths (in Excel units, approximate)
    col_widths = []
    for col in range(1, max_col_used + 1):
        width = ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width
        col_widths.append(width if width else 8)

    wb.close()

    return {
        "max_row": max_row,
        "max_col": max_col,
        "used_rows": (min_row, max_row_used),
        "used_cols": (min_col, max_col_used),
        "col_widths": col_widths,
        "total_width_units": sum(col_widths),
    }


def convert_xlsx_to_pdf_quality(file_path: str, dpi: int = 300) -> str:
    """Convert XLSX to PDF with quality analysis"""
    from app.services.converter import DocumentConverter

    log_step("1", "Initializing converter...")
    converter = DocumentConverter(temp_dir="/tmp/converter_test")

    log_step("2", "Reading XLSX file...")
    with open(file_path, "rb") as f:
        xlsx_data = f.read()
    log_step("2", f"Input size: {len(xlsx_data)} bytes")

    log_step("3", "Analyzing XLSX dimensions...")
    dims = analyze_xlsx_dimensions(file_path)
    print(f"   - Used rows: {dims['used_rows']}")
    print(f"   - Used cols: {dims['used_cols']}")
    print(f"   - Total column width units: {dims['total_width_units']}")

    log_step("4", "Converting to PDF (DPI: {})...".format(dpi))
    pdf_data = converter.excel_to_pdf(xlsx_data)
    log_step("4", f"Output size: {len(pdf_data)} bytes")

    # Save output
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_pdf = os.path.join(OUTPUT_DIR, "xlsx_quality_output.pdf")
    with open(output_pdf, "wb") as f:
        f.write(pdf_data)
    log_step("4", f"Saved to: {output_pdf}")

    return output_pdf


def analyze_pdf_quality(pdf_path: str) -> dict:
    """Analyze PDF to check for quality issues"""
    import fitz

    log_step("5", "Analyzing PDF quality...")

    doc = fitz.open(pdf_path)

    results = {"pages": len(doc), "page_sizes": [], "content_check": [], "issues": []}

    for page_num in range(len(doc)):
        page = doc[page_num]

        # Get page dimensions (in points)
        rect = page.rect
        page_width = rect.width
        page_height = rect.height

        results["page_sizes"].append(
            {
                "width": page_width,
                "height": page_height,
                "width_mm": page_width * 0.352778,
                "height_mm": page_height * 0.352778,
            }
        )

        # Check text content
        text = page.get_text()
        text_length = len(text.strip())

        # Check images
        images = page.get_images()

        # Check for content that might be cut off
        # (simple heuristic: check if there's content near edges)
        clip_margin = 20  # points
        has_content_near_edge = False

        # Get all text blocks
        blocks = page.get_text("blocks")
        for block in blocks:
            block_rect = fitz.Rect(block[:4])
            if (
                block_rect.x0 < clip_margin
                or block_rect.y0 < clip_margin
                or page_width - block_rect.x1 < clip_margin
                or page_height - block_rect.y1 < clip_margin
            ):
                has_content_near_edge = True
                break

        results["content_check"].append(
            {
                "page": page_num + 1,
                "text_length": text_length,
                "images": len(images),
                "content_near_edge": has_content_near_edge,
            }
        )

        # Check for potential issues
        if text_length < 10:
            results["issues"].append(f"Page {page_num + 1}: Very little text content")

        if len(images) == 0 and text_length < 50:
            results["issues"].append(
                f"Page {page_num + 1}: May be empty or have rendering issues"
            )

    doc.close()

    return results


async def capture_pdf_preview(pdf_path: str) -> bool:
    """Capture PDF preview screenshot"""
    from playwright.async_api import async_playwright

    log_step("6", "Capturing PDF preview screenshot...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()

        # Set viewport for PDF viewing
        await page.set_viewport_size({"width": 1200, "height": 1600})

        # Open PDF
        pdf_url = f"file://{os.path.abspath(pdf_path)}"
        await page.goto(pdf_url)
        await page.wait_for_timeout(3000)

        # Take screenshot
        screenshot_path = os.path.join(OUTPUT_DIR, "xlsx_quality_preview.png")
        await page.screenshot(path=screenshot_path, full_page=False)

        await browser.close()

        log_step("6", f"Screenshot saved: {screenshot_path}")
        return True


def check_xlsx_quality_issues(dims: dict, pdf_results: dict) -> list:
    """Identify quality issues and suggest fixes"""
    issues = []

    # Check 1: Total column width vs PDF page width
    # A4 width at 72 DPI is approximately 595 points
    # At 300 DPI, it's much larger
    pdf_width = (
        pdf_results["page_sizes"][0]["width"] if pdf_results["page_sizes"] else 595
    )

    # Excel column width units to PDF points conversion is approximate
    # Standard Excel column width ~8 units = ~72 points at 100% zoom
    estimated_total_width = dims["total_width_units"] * 9  # Approximate conversion

    if estimated_total_width > pdf_width * 0.9:
        issues.append(
            {
                "type": "content_truncation",
                "severity": "high",
                "message": f"Content may be truncated: estimated width {estimated_total_width} > PDF width {pdf_width}",
                "suggestion": "Consider using landscape orientation or reducing content",
            }
        )

    # Check 2: Many columns may not fit
    if dims["used_cols"][1] - dims["used_cols"][0] > 8:
        issues.append(
            {
                "type": "many_columns",
                "severity": "medium",
                "message": f"Many columns ({dims['used_cols'][1] - dims['used_cols'][0] + 1}) may not fit on single page",
                "suggestion": "Consider auto-sizing or landscape mode",
            }
        )

    # Check 3: Check for near-edge content (possible cutoff)
    for check in pdf_results["content_check"]:
        if check.get("content_near_edge"):
            issues.append(
                {
                    "type": "edge_content",
                    "severity": "low",
                    "message": f"Page {check['page']}: Content near edge may be cut off",
                    "suggestion": "Increase page margins",
                }
            )

    return issues


async def main():
    """Main quality test runner"""
    print("=" * 70)
    print("XLSX to PDF Quality Test")
    print("=" * 70)
    print()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Step 1-4: Convert
    try:
        pdf_path = convert_xlsx_to_pdf_quality(XLSX_FILE)
    except Exception as e:
        print(f"\n✗ Conversion failed: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)

    # Step 5: Analyze
    pdf_results = analyze_pdf_quality(pdf_path)

    print(f"\n[5] PDF Analysis Results:")
    print(f"    Pages: {pdf_results['pages']}")
    for check in pdf_results["content_check"]:
        print(
            f"    Page {check['page']}: {check['text_length']} chars, {check['images']} images"
        )

    if pdf_results["issues"]:
        print(f"\n    Issues found: {len(pdf_results['issues'])}")
        for issue in pdf_results["issues"]:
            print(f"      - {issue}")

    # Analyze XLSX dimensions
    xlsx_dims = analyze_xlsx_dimensions(XLSX_FILE)

    # Check for quality issues
    issues = check_xlsx_quality_issues(xlsx_dims, pdf_results)

    if issues:
        print(f"\n[6] Quality Issues Detected:")
        for issue in issues:
            print(f"    [{issue['severity'].upper()}] {issue['message']}")
            print(f"        → {issue['suggestion']}")
    else:
        print(f"\n[6] No quality issues detected")

    # Step 6: Preview screenshot
    await capture_pdf_preview(pdf_path)

    print(f"\n" + "=" * 70)
    print(f"Quality test complete. Check output in: {OUTPUT_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    asyncio.run(main())
